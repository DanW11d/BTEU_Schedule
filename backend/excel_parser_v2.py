"""
Парсер Excel файлов с расписанием для BTEU Schedule (версия 2)
Поддерживает столбцовый формат: каждый столбец = группа
"""
import re
import os
from typing import List, Dict, Optional, Tuple
from openpyxl import load_workbook, Workbook
from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet


class ExcelScheduleParserV2:
    """Парсер Excel файлов с расписанием (столбцовый формат)"""
    
    DAYS_OF_WEEK = {
        'понедельник': 1,
        'вторник': 2,
        'среда': 3,
        'четверг': 4,
        'пятница': 5,
        'суббота': 6
    }
    
    def __init__(self, file_path: str):
        self.file_path = file_path
        self.workbook = None
        self.worksheet = None
        
    def load_file(self) -> None:
        """Загружает Excel файл (поддерживает .xlsx и .xls)"""
        try:
            file_ext = os.path.splitext(self.file_path)[1].lower()
            
            if file_ext == '.xls':
                try:
                    import xlrd
                    xls_book = xlrd.open_workbook(self.file_path)
                    xls_sheet = xls_book.sheet_by_index(0)
                    
                    temp_wb = Workbook()
                    temp_ws = temp_wb.active
                    
                    for row_idx in range(xls_sheet.nrows):
                        for col_idx in range(xls_sheet.ncols):
                            cell_value = xls_sheet.cell_value(row_idx, col_idx)
                            temp_ws.cell(row=row_idx + 1, column=col_idx + 1, value=cell_value)
                    
                    self.workbook = temp_wb
                    self.worksheet = temp_ws
                except ImportError:
                    raise ValueError("Для чтения .xls файлов требуется библиотека xlrd")
                except Exception as e:
                    raise ValueError(f"Ошибка чтения .xls файла: {str(e)}")
            else:
                self.workbook = load_workbook(self.file_path, data_only=True)
                self.worksheet = self.workbook.active
        except Exception as e:
            raise ValueError(f"Ошибка загрузки файла: {str(e)}")
    
    def parse(self, group_code: str, group_id: int) -> List[Dict]:
        """
        Парсит расписание из Excel файла (столбцовый формат)
        
        Args:
            group_code: Код группы (например, "П-1")
            group_id: ID группы в БД
        """
        if not self.worksheet:
            self.load_file()
        
        lessons = []
        
        # Находим столбец с нужной группой
        group_col = self._find_group_column(group_code)
        if group_col is None:
            return lessons  # Группа не найдена в файле
        
        # Находим строку с заголовком (где указаны группы)
        header_row = self._find_header_row()
        if header_row is None:
            return lessons
        
        # Парсим данные начиная со следующей строки после заголовка
        current_day = None
        current_time = None
        current_lesson_number = 0
        accumulated_cell_content = None  # Накопленное содержимое для многострочных предметов
        accumulated_lesson_number = None  # Номер пары для накопленного содержимого
        current_week_parity = None  # Текущая четность недели (определяется по разделителям)
        day_separator_count = {}  # Счетчик разделителей для каждого дня
        
        for row_idx in range(header_row + 1, self.worksheet.max_row + 1):
            day_cell = self.worksheet.cell(row=row_idx, column=1).value
            time_cell = self.worksheet.cell(row=row_idx, column=2).value
            group_cell = self.worksheet.cell(row=row_idx, column=group_col).value
            
            # Определяем день недели
            if day_cell:
                day_str = str(day_cell).strip().lower()
                if day_str in self.DAYS_OF_WEEK:
                    # Если был накопленный контент - обрабатываем его перед сменой дня
                    if accumulated_cell_content and current_day and current_time and accumulated_lesson_number:
                        # Используем текущую четность из разделителей, или odd по умолчанию (семестр начинается с нечетной)
                        week_parity_to_use = current_week_parity if current_week_parity is not None else 'odd'
                        lesson = self._parse_lesson_cell(
                            accumulated_cell_content,
                            group_id,
                            group_code,
                            current_day,
                            current_time,
                            accumulated_lesson_number,
                            week_parity_to_use
                        )
                        if lesson:
                            lessons.append(lesson)
                    accumulated_cell_content = None
                    accumulated_lesson_number = None
                    current_day = self.DAYS_OF_WEEK[day_str]
                    current_lesson_number = 0
                    # Сбрасываем счетчик разделителей для нового дня
                    if current_day not in day_separator_count:
                        day_separator_count[current_day] = 0
                    # НЕ сбрасываем четность при смене дня - она сохраняется между днями
                    # Если четность не установлена, будет использована нечетная по умолчанию
            
            # Определяем время
            new_time_found = False
            if time_cell:
                time_str = str(time_cell).strip()
                if self._is_time_format(time_str):
                    new_time_found = True
                    # Если есть накопленное содержимое, проверяем, не является ли следующая ячейка продолжением
                    # Если нет накопленного содержимого или следующая ячейка не продолжение - обрабатываем накопленное
                    # Но мы еще не знаем, что будет в следующей ячейке, поэтому сохраняем накопленное
                    # и обработаем его, когда встретим следующую ячейку группы
                    current_time = time_str
                    # Определяем номер пары по времени (а не инкрементально!)
                    current_lesson_number = self._get_lesson_number_from_time(time_str)
                    # Защита от некорректных значений
                    if current_lesson_number < 1:
                        current_lesson_number = 1
                    if current_lesson_number > 7:
                        current_lesson_number = 7
                    # НЕ меняем accumulated_lesson_number - он останется прежним для возможного объединения
                    # НО если накопленное содержимое не будет объединено со следующей ячейкой,
                    # оно будет обработано при встрече следующей ячейки группы
            
            # Проверяем на разделитель недель (строка с дефисами/тире)
            if group_cell:
                cell_str = str(group_cell).strip()
                # Разделитель недель - строка, состоящая в основном из дефисов/тире
                if cell_str and (cell_str.startswith('—') or cell_str.startswith('─') or 
                                 (len(cell_str.replace(' ', '').replace('—', '').replace('─', '').replace('-', '')) < 3 and 
                                  ('—' in cell_str or '─' in cell_str or cell_str.count('-') > 5))):
                    # Это разделитель недель - переключаем четность
                    # Логика: первый разделитель после дня означает начало блока занятий
                    # После первого разделителя идет НЕЧЕТНАЯ неделя (семестр начинается с нечетной, 1.09)
                    # После второго разделителя идет ЧЕТНАЯ неделя
                    # И так далее чередуются
                    # ВАЖНО: если в дне только 1 разделитель, все занятия остаются нечетной недели
                    # Четная неделя создается только если есть 2+ разделителя в одном дне
                    if current_day:
                        day_separator_count[current_day] = day_separator_count.get(current_day, 0) + 1
                        separator_count_for_day = day_separator_count[current_day]
                        
                        if separator_count_for_day == 1:
                            # Первый разделитель - устанавливаем четную (ИНВЕРТИРОВАНО)
                            current_week_parity = 'even'
                        elif separator_count_for_day == 2:
                            # Второй разделитель - переключаем на нечетную (есть явное разделение)
                            current_week_parity = 'odd'
                        else:
                            # Третий и далее - чередуем
                            if current_week_parity == 'even':
                                current_week_parity = 'odd'
                            else:
                                current_week_parity = 'even'
                    else:
                        # Если день не определен, используем общую логику
                        # Начинаем с нечетной (семестр начинается с нечетной недели)
                        if current_week_parity is None:
                            current_week_parity = 'odd'
                        elif current_week_parity == 'odd':
                            current_week_parity = 'even'
                        else:
                            current_week_parity = 'odd'
                    # Пропускаем разделитель
                    continue
            
            # Обрабатываем ячейку группы
            if group_cell and current_day and current_time:
                cell_str = str(group_cell).strip()
                
                # Пропускаем пустые ячейки
                if not cell_str or cell_str.startswith('—') or cell_str.startswith('-'):
                    # Если была накопленная ячейка - обрабатываем её
                    if accumulated_cell_content and accumulated_lesson_number:
                        # Используем текущую четность из разделителей, или odd по умолчанию (семестр начинается с нечетной)
                        week_parity_to_use = current_week_parity if current_week_parity is not None else 'odd'
                        lesson = self._parse_lesson_cell(
                            accumulated_cell_content,
                            group_id,
                            group_code,
                            current_day,
                            current_time,
                            accumulated_lesson_number,
                            week_parity_to_use
                        )
                        if lesson:
                            lessons.append(lesson)
                        accumulated_cell_content = None
                        accumulated_lesson_number = None
                    continue
                
                # Проверяем, является ли текущая ячейка явным продолжением предыдущего накопленного содержимого
                # (например, "ТЕХНОЛОГИЯХ)" после "(В ИНФОРМАЦИОННЫХ")
                is_explicit_continuation = False
                if accumulated_cell_content and cell_str:
                    cell_upper = cell_str.strip().upper()
                    prev_upper = accumulated_cell_content.strip().upper()
                    # Проверяем явные паттерны продолжения
                    # Убираем лишние пробелы и символы для сравнения
                    cell_clean = cell_upper.replace('"', '').replace("'", '').strip()
                    prev_clean = prev_upper.replace('"', '').replace("'", '').strip()
                    
                    is_explicit_continuation = (
                        (cell_clean.startswith('ТЕХНОЛОГИЯХ') or 
                         cell_clean.startswith('ТЕХНОЛОГИЯ') or
                         (cell_clean.endswith(')') and len(cell_clean) < 20)) and
                        (prev_clean.endswith('(В ИНФОРМАЦИОННЫХ') or
                         prev_clean.endswith('(В ИНФОРМАЦИОННЫХ ') or
                         prev_clean.endswith('ИНФОРМАЦИОННЫХ') or
                         prev_clean.endswith('ИНФОРМАЦИОННЫХ ') or
                         '(В ИНФОРМАЦИОННЫХ' in prev_clean)
                    )
                
                # Определяем, является ли это продолжением предыдущей ячейки
                is_likely_continuation = (
                    cell_str and len(cell_str) > 0 and (
                        cell_str[0].islower() or  # Начинается с маленькой буквы
                        cell_str[0] in '()' or  # Начинается со скобки
                        (len(cell_str) < 30 and not any(c.isupper() for c in cell_str[:3]))  # Короткая и без заглавных в начале
                    )
                )
                
                # Проверяем, заканчивается ли предыдущее содержимое на незавершенное слово/скобку
                prev_ends_unfinished = False
                if accumulated_cell_content:
                    prev_trimmed = accumulated_cell_content.strip()
                    # Заканчивается на открывающую скобку, дефис, или незавершенное слово
                    prev_ends_unfinished = (
                        prev_trimmed.endswith('(') or
                        prev_trimmed.endswith('(В') or
                        prev_trimmed.endswith('(В ИНФОРМАЦИОННЫХ') or
                        prev_trimmed.endswith('(В ИНФОРМАЦИОННЫХ ') or
                        prev_trimmed.endswith('-') or
                        (len(prev_trimmed) > 50 and not prev_trimmed.endswith('.') and not prev_trimmed.endswith(')'))
                    )
                
                # Проверяем, является ли это продолжением
                # Если явное продолжение - объединяем всегда (даже если номер пары изменился)
                # Если обычное продолжение - только если нет нового дня
                is_continuation = False
                if accumulated_cell_content and accumulated_lesson_number:
                    if is_explicit_continuation:
                        # Явное продолжение - объединяем всегда (даже если номер пары изменился)
                        is_continuation = True
                    elif not day_cell and (is_likely_continuation or prev_ends_unfinished):
                        # Обычное продолжение - только если нет нового дня
                        is_continuation = True
                
                if is_continuation:
                    # Объединяем с предыдущим содержимым, используя номер пары предыдущего
                    # Убираем лишние пробелы при объединении
                    accumulated_cell_content = (accumulated_cell_content.strip() + " " + cell_str.strip()).strip()
                    # Номер пары остается прежним (из предыдущего накопления)
                else:
                    # Это новая ячейка - обрабатываем предыдущую, если была
                    # ВАЖНО: обрабатываем накопленное содержимое ТОЛЬКО если оно не пустое
                    # и если это действительно новая ячейка (не продолжение)
                    if accumulated_cell_content and accumulated_lesson_number:
                        # Проверяем, не является ли текущая ячейка продолжением предыдущей
                        # Если нет - обрабатываем накопленное
                        # Используем текущую четность из разделителей, или odd по умолчанию (семестр начинается с нечетной)
                        week_parity_to_use = current_week_parity if current_week_parity is not None else 'odd'
                        lesson = self._parse_lesson_cell(
                            accumulated_cell_content,
                            group_id,
                            group_code,
                            current_day,
                            current_time,
                            accumulated_lesson_number,
                            week_parity_to_use
                        )
                        if lesson:
                            lessons.append(lesson)
                        # Очищаем накопленное содержимое после обработки
                        accumulated_cell_content = None
                        accumulated_lesson_number = None
                    # Начинаем накапливать новое содержимое
                    accumulated_cell_content = cell_str
                    accumulated_lesson_number = current_lesson_number
        
        # Обрабатываем последнюю накопленную ячейку
        if accumulated_cell_content and current_day and current_time and accumulated_lesson_number:
            # Используем текущую четность из разделителей, или odd по умолчанию
            week_parity_to_use = current_week_parity if current_week_parity is not None else 'odd'
            lesson = self._parse_lesson_cell(
                accumulated_cell_content,
                group_id,
                group_code,
                current_day,
                current_time,
                accumulated_lesson_number,
                week_parity_to_use
            )
            if lesson:
                lessons.append(lesson)
        
        return lessons
    
    def _find_group_column(self, group_code: str) -> Optional[int]:
        """Находит столбец с указанной группой"""
        # Ищем в первых 20 строках заголовок с группами
        for row_idx in range(1, min(20, self.worksheet.max_row + 1)):
            for col_idx in range(1, min(20, self.worksheet.max_column + 1)):
                cell = self.worksheet.cell(row=row_idx, column=col_idx)
                if cell.value:
                    cell_str = str(cell.value).strip()
                    # Ищем код группы в ячейке (может быть "П -11", "П-1", "П-11" и т.д.)
                    if self._matches_group_code(cell_str, group_code):
                        return col_idx
        return None
    
    def _matches_group_code(self, cell_str: str, group_code: str) -> bool:
        """Проверяет, соответствует ли ячейка коду группы"""
        # Нормализуем строки
        cell_upper = cell_str.upper().strip()
        group_upper = group_code.upper().strip()
        
        # Убираем пробелы и дефисы для сравнения
        cell_normalized = re.sub(r'[\s\-]+', '', cell_upper)
        group_normalized = re.sub(r'[\s\-]+', '', group_upper)
        
        # Проверяем точное совпадение
        if cell_normalized == group_normalized:
            return True
        
        # Проверяем, содержит ли ячейка код группы
        # Например, "П -11" содержит "П" и "11", "П-1" содержит "П" и "1"
        group_letter = re.sub(r'[^А-Я]', '', group_normalized)
        group_number = re.sub(r'[^0-9]', '', group_normalized)
        
        cell_letter = re.sub(r'[^А-Я]', '', cell_normalized)
        cell_number = re.sub(r'[^0-9]', '', cell_normalized)
        
        # Если буква совпадает и номер начинается с нужного
        if group_letter and cell_letter == group_letter:
            if group_number and cell_number.startswith(group_number):
                return True
            # Также проверяем обратное - может быть "П-1" в файле "П-11"
            if cell_number and group_number.startswith(cell_number):
                return True
        
        # Проверяем паттерны типа "П -11", "П-1" и т.д.
        # Создаем паттерн для поиска группы с возможными пробелами и дефисами
        # Например, "П-1" -> "П\s*-?\s*1"
        pattern_parts = []
        for char in group_code:
            if char.isalpha():
                pattern_parts.append(char)
                pattern_parts.append(r'\s*-?\s*')
            else:
                pattern_parts.append(re.escape(char))
        pattern = ''.join(pattern_parts)
        if re.search(pattern, cell_str, re.IGNORECASE):
            return True
        
        return False
    
    def _find_header_row(self) -> Optional[int]:
        """Находит строку с заголовком (где указаны группы)"""
        for row_idx in range(1, min(20, self.worksheet.max_row + 1)):
            row_str = " ".join([
                str(self.worksheet.cell(row=row_idx, column=col).value or "")
                for col in range(1, min(10, self.worksheet.max_column + 1))
            ]).lower()
            
            # Ищем строку с заголовком типа "Дни | Время | П -11 | П -12"
            if "дни" in row_str and "время" in row_str:
                return row_idx
        
        return None
    
    def _get_lesson_number_from_time(self, time_str: str) -> int:
        """Определяет номер пары по времени занятия"""
        time_str_lower = time_str.lower().replace(' ', '')
        
        # Соответствие времени и номера пары
        # Пара 1: 9:00-9:45 и 9:50-10:35
        if any(t in time_str_lower for t in ['9:00-9:45', '9.00-9.45', '9:50-10:35', '9.50-10.35']):
            return 1
        # Пара 2: 10:50-11:35 и 11:40-12:25
        elif any(t in time_str_lower for t in ['10:50-11:35', '10.50-11.35', '11:40-12:25', '11.40-12.25']):
            return 2
        # Пара 3: 12:55-13:40 и 13:45-14:30
        elif any(t in time_str_lower for t in ['12:55-13:40', '12.55-13.40', '13:45-14:30', '13.45-14.30']):
            return 3
        # Пара 4: 14:40-15:25 и 15:30-16:15
        elif any(t in time_str_lower for t in ['14:40-15:25', '14.40-15.25', '15:30-16:15', '15.30-16.15']):
            return 4
        # Пара 5: 16:25-17:10 и 17:15-18:00
        elif any(t in time_str_lower for t in ['16:25-17:10', '16.25-17.10', '17:15-18:00', '17.15-18.00']):
            return 5
        # Пара 6: 18:10-18:55 и 19:00-19:45 (если есть)
        elif any(t in time_str_lower for t in ['18:10-18:55', '18.10-18.55', '19:00-19:45', '19.00-19.45']):
            return 6
        # Пара 7: 19:50-20:35 и 20:40-21:25 (если есть)
        elif any(t in time_str_lower for t in ['19:50-20:35', '19.50-20.35', '20:40-21:25', '20.40-21.25']):
            return 7
        else:
            # Если время не распознано, возвращаем 0 (будет использовано предыдущее значение)
            return 0
    
    def _is_time_format(self, time_str: str) -> bool:
        """Проверяет, является ли строка временем"""
        # Форматы: "9.00-9.45", "9:00-9:45", "09:00-10:35"
        time_pattern = r'\d{1,2}[.:]\d{2}\s*-\s*\d{1,2}[.:]\d{2}'
        return bool(re.search(time_pattern, time_str))
    
    def _parse_lesson_cell(
        self,
        cell_content: str,
        group_id: int,
        group_code: str,
        day_of_week: int,
        time: str,
        lesson_number: int,
        week_parity_override: str = None
    ) -> Optional[Dict]:
        """Парсит ячейку с занятием"""
        cell_content = cell_content.strip()
        
        # Пропускаем пустые ячейки и разделители
        if not cell_content or cell_content.startswith('—') or cell_content.startswith('-'):
            return None
        
        # Извлекаем четность недели
        # Приоритет: 1) override из разделителей, 2) из содержимого ячейки, 3) odd (нечетная)
        # По умолчанию используем нечетную неделю (семестр начинается с нечетной, 1.09)
        if week_parity_override:
            week_parity = week_parity_override
        else:
            extracted = self._extract_week_parity(cell_content)
            if extracted is None:
                # Если не удалось определить из содержимого, используем odd (нечетная)
                week_parity = 'odd'
            else:
                week_parity = extracted
        
        # Извлекаем предмет, преподавателя, аудиторию
        subject = self._extract_subject(cell_content)
        teacher = self._extract_teacher(cell_content)
        classroom = self._extract_classroom(cell_content)
        lesson_type = self._detect_lesson_type(cell_content)
        
        # Проверяем, что есть предмет (не только преподаватель)
        # Если subject пустой или слишком короткий, и есть только преподаватель - пропускаем
        if not subject or len(subject.strip()) < 3:
            # Проверяем, может быть это только преподаватель без предмета
            if teacher and not subject:
                return None  # Пропускаем записи, где только преподаватель
            if not subject:
                return None
        
        # Защита от некорректных значений lesson_number
        valid_lesson_number = lesson_number
        if valid_lesson_number < 1:
            valid_lesson_number = 1
        if valid_lesson_number > 7:
            valid_lesson_number = 7
        
        return {
            'group_id': group_id,
            'day_of_week': day_of_week,
            'lesson_number': valid_lesson_number,
            'subject': subject,
            'teacher': teacher or '',
            'classroom': classroom or '',
            'lesson_type': lesson_type,
            'week_parity': week_parity,
            'building': None,
            'notes': None
        }
    
    def _extract_week_parity(self, content: str) -> Optional[str]:
        """Извлекает четность недели из содержимого"""
        # Сначала проверяем на ключевые слова (более надежно)
        content_lower = content.lower()
        # ИНВЕРТИРОВАНО: нечет → even, чет → odd
        if any(word in content_lower for word in ['нечет', 'подчеты', 'нечёт']):
            return 'even'
        if any(word in content_lower for word in ['чет', 'одцчеты', 'чёт', 'четн']):
            return 'odd'
        
        # Ищем паттерны типа "1-13", "1-28", "2-6"
        # ВАЖНО: диапазон "X-Y" означает недели с X по Y
        # Если диапазон большой (например, 3-18 или 2-22), это означает, что занятие идет
        # и на четной, и на нечетной неделе в этом диапазоне
        # Поэтому НЕ определяем четность по диапазонам - она должна определяться разделителями
        week_match = re.search(r'(\d+)\s*-\s*(\d+)', content)
        if week_match:
            start_week = int(week_match.group(1))
            end_week = int(week_match.group(2))
            # Если диапазон очень маленький (1 неделя), можно определить по началу (ИНВЕРТИРОВАНО)
            if end_week == start_week:
                if start_week % 2 == 1:
                    return 'even'  # ИНВЕРТИРОВАНО: нечетная неделя → even
                else:
                    return 'odd'   # ИНВЕРТИРОВАНО: четная неделя → odd
            # Для больших диапазонов возвращаем None - четность должна определяться разделителями
            return None
        
        # Если не нашли явных указаний, возвращаем None
        # Четность будет определена из разделителей
        return None
    
    def _extract_subject(self, content: str) -> str:
        """Извлекает название предмета"""
        # Убираем номера недель, преподавателя, аудиторию
        subject = content
        
        # Убираем префиксы типа "с 22.09.25"
        subject = re.sub(r'с\s+\d{2}\.\d{2}\.\d{2,4}\s+', '', subject, flags=re.IGNORECASE)
        
        # Убираем номера недель (например, "1-13", "1-28")
        subject = re.sub(r'\d+\s*-\s*\d+', '', subject)
        
        # Убираем преподавателя (паттерны: "доц.Иванов И.И.", "ст.пр.Петров П.П.", "пр.Сидоров С.С.")
        # Ищем и удаляем преподавателя в конце строки, но сохраняем закрывающие скобки
        # Если есть закрывающая скобка перед преподавателем, удаляем только преподавателя после неё
        # Паттерн: "ТЕКСТ)" доц.Иванов -> "ТЕКСТ)"
        subject = re.sub(r'\)\s*[.,]?\s*([а-яё]+\.\s*[А-ЯЁ][а-яё]+\s+[А-ЯЁ]\.[А-ЯЁ]\.).*$', ')', subject, flags=re.IGNORECASE)
        # Если преподаватель без закрывающей скобки перед ним, удаляем его и всё после
        subject = re.sub(r'([^)])\s*([а-яё]+\.\s*[А-ЯЁ][а-яё]+\s+[А-ЯЁ]\.[А-ЯЁ]\.).*$', r'\1', subject, flags=re.IGNORECASE)
        # Также удаляем, если преподаватель стоит после запятой/точки
        subject = re.sub(r'[.,]\s*([а-яё]+\.\s*[А-ЯЁ][а-яё]+\s+[А-ЯЁ]\.[А-ЯЁ]\.).*$', '', subject, flags=re.IGNORECASE)
        
        # Убираем аудиторию (паттерны: "3-11", "5-41", "Ауд. 2-22")
        subject = re.sub(r'[а-яё]*\s*ауд\.?\s*', '', subject, flags=re.IGNORECASE)
        subject = re.sub(r'\d+[-\s]*\d*[а-яё]*\s*$', '', subject, flags=re.IGNORECASE)
        
        # Убираем лишние пробелы, но сохраняем закрывающие скобки, если они часть названия
        # Убираем только одиночные скобки в начале/конце, но не закрывающие скобки после текста
        # Например: "ПРЕДМЕТ (В ТЕХНОЛОГИЯХ)" - скобки часть названия, не удаляем
        # Но: "(ПРЕДМЕТ)" - одиночные скобки, можно удалить
        if subject.count('(') == subject.count(')'):
            # Сбалансированные скобки - это часть названия, не удаляем
            pass
        else:
            # Несбалансированные - убираем одиночные в начале/конце
            subject = re.sub(r'^\s*\(+', '', subject)
            if not subject.endswith(')'):
                subject = re.sub(r'\s*\)+$', '', subject)
        
        return subject.strip()
    
    def _extract_teacher(self, content: str) -> Optional[str]:
        """Извлекает имя преподавателя"""
        # Ищем паттерны типа "доц.Иванов И.И.", "ст.пр.Петров П.П."
        teacher_match = re.search(r'([а-яё]+\.\s*[А-ЯЁ][а-яё]+\s+[А-ЯЁ]\.[А-ЯЁ]\.)', content, re.IGNORECASE)
        if teacher_match:
            return teacher_match.group(1).strip()
        return None
    
    def _extract_classroom(self, content: str) -> Optional[str]:
        """Извлекает аудиторию"""
        # Ищем паттерны типа "3-11", "5-41", "Бол. Акт. Зал"
        classroom_match = re.search(r'(\d+[-\s]*\d*[а-яё]*|[А-ЯЁ][а-яё]+\s+[А-ЯЁ][а-яё]+\s+[А-ЯЁ][а-яё]+)', content, re.IGNORECASE)
        if classroom_match:
            return classroom_match.group(1).strip()
        return None
    
    def _detect_lesson_type(self, content: str) -> str:
        """Определяет тип занятия по регистру букв в названии предмета"""
        # Сначала проверяем явные указания в тексте
        content_lower = content.lower()
        
        if any(word in content_lower for word in ['лекция', 'лекц']):
            return 'lecture'
        if any(word in content_lower for word in ['практика', 'практ', 'пр.']):
            return 'practice'
        if any(word in content_lower for word in ['лабораторная', 'лаб']):
            return 'laboratory'
        
        # Извлекаем название предмета (без преподавателя, аудитории и т.д.)
        subject = self._extract_subject(content)
        
        if not subject:
            return 'lecture'  # По умолчанию, если не удалось извлечь
        
        # Подсчитываем количество заглавных и строчных букв (только кириллица и латиница)
        uppercase_count = 0
        lowercase_count = 0
        total_letters = 0
        
        for char in subject:
            if char.isalpha():
                total_letters += 1
                if char.isupper():
                    uppercase_count += 1
                elif char.islower():
                    lowercase_count += 1
        
        # Если нет букв, возвращаем лекцию по умолчанию
        if total_letters == 0:
            return 'lecture'
        
        # Вычисляем процент заглавных букв
        uppercase_percentage = (uppercase_count / total_letters) * 100
        
        # ИНВЕРТИРОВАННАЯ ЛОГИКА: КАПС (большие буквы) → практика, маленькие → лекция
        # Если больше 50% букв заглавные (КАПС) → практика
        if uppercase_percentage > 50:
            return 'practice'
        
        # Если больше 20% оставшихся букв (кроме первой) заглавные → практика
        if total_letters > 1:
            remaining_uppercase = max(0, uppercase_count - 1) if subject[0].isupper() else uppercase_count
            remaining_total = total_letters - 1
            if remaining_total > 0:
                remaining_uppercase_percentage = (remaining_uppercase / remaining_total) * 100
                if remaining_uppercase_percentage > 20:
                    return 'practice'
        
        # Иначе (в основном строчные буквы) → лекция
        return 'lecture'
    
    def validate_lessons(self, lessons: List[Dict]) -> Tuple[bool, List[str]]:
        """Валидирует список занятий"""
        errors = []
        
        for lesson in lessons:
            if not lesson.get('subject'):
                errors.append(f"Занятие без предмета: {lesson}")
            if not lesson.get('day_of_week') or not (1 <= lesson['day_of_week'] <= 6):
                errors.append(f"Некорректный день недели: {lesson}")
            if not lesson.get('lesson_number') or not (1 <= lesson['lesson_number'] <= 7):
                errors.append(f"Некорректный номер пары: {lesson}")
        
        return len(errors) == 0, errors

