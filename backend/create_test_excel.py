"""
Создает тестовый Excel файл для проверки парсера
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

def create_test_schedule():
    """Создает тестовый Excel файл с расписанием"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Расписание"
    
    # Шапка таблицы
    ws['A1'] = "РАСПИСАНИЕ ЗАНЯТИЙ"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:F1')
    
    ws['A2'] = "Группа: П-1"
    ws['A2'].font = Font(bold=True)
    
    # Заголовки столбцов
    headers = ["№", "День недели", "Предмет", "Преподаватель", "Аудитория", "Примечания"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    row = 4
    
    # Разделитель для четной недели
    ws.cell(row=row, column=1).value = "───────── ОДЦ ЧЕТЫ ─────────"
    ws.merge_cells(f'A{row}:F{row}')
    row += 1
    
    # Четная неделя - Понедельник
    lessons_even_monday = [
        [1, "Понедельник", "БУХГАЛТЕРСКИЙ УЧЕТ", "доц. Иванов И.И.", "3-11", ""],
        [2, "Понедельник", "Экономика предприятия", "проф. Петров П.П.", "5-41", ""],
    ]
    
    for lesson in lessons_even_monday:
        for col, value in enumerate(lesson, 1):
            ws.cell(row=row, column=col).value = value
        row += 1
    
    # Четная неделя - Вторник
    ws.cell(row=row, column=1).value = ""
    row += 1
    lessons_even_tuesday = [
        [1, "Вторник", "Финансовый менеджмент", "доц. Сидоров С.С.", "2-15", ""],
    ]
    
    for lesson in lessons_even_tuesday:
        for col, value in enumerate(lesson, 1):
            ws.cell(row=row, column=col).value = value
        row += 1
    
    # Разделитель для нечетной недели
    row += 1
    ws.cell(row=row, column=1).value = "───────── ПОД ЧЕТЫ ─────────"
    ws.merge_cells(f'A{row}:F{row}')
    row += 1
    
    # Нечетная неделя - Понедельник
    lessons_odd_monday = [
        [1, "Понедельник", "Международные стандарты аудита", "доц. Томалева Е.Г.", "3-11", ""],
        [2, "Понедельник", "Внутренний аудит", "доц. Томалева Е.Г.", "5-41", ""],
    ]
    
    for lesson in lessons_odd_monday:
        for col, value in enumerate(lesson, 1):
            ws.cell(row=row, column=col).value = value
        row += 1
    
    # Нечетная неделя - Среда
    row += 1
    lessons_odd_wednesday = [
        [1, "Среда", "Налогообложение", "доц. Козлова К.К.", "3-20", ""],
    ]
    
    for lesson in lessons_odd_wednesday:
        for col, value in enumerate(lesson, 1):
            ws.cell(row=row, column=col).value = value
        row += 1
    
    # Подвал
    row += 2
    ws.cell(row=row, column=1).value = "Расписание составлено автоматически"
    ws.cell(row=row, column=1).font = Font(italic=True)
    
    # Настройка ширины столбцов
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 20
    
    # Сохраняем файл
    filename = "test_schedule.xlsx"
    wb.save(filename)
    print(f"[OK] Test file created: {filename}")
    print(f"  Path: {os.path.abspath(filename)}")
    return filename

if __name__ == '__main__':
    import os
    import sys
    # Устанавливаем UTF-8 для вывода
    if sys.platform == 'win32':
        import io
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
    
    filename = create_test_schedule()
    print(f"\nNow you can test the parser:")
    print(f"  python test_excel_parser.py {filename} P-1")
    print(f"\nOr upload via API:")
    print(f"  curl -X POST http://localhost:8000/v1/admin/parse-excel -F \"file=@{filename}\" -F \"group_code=P-1\"")

