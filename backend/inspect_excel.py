"""
Скрипт для проверки структуры Excel файла
Помогает понять формат данных для улучшения парсера
"""
import sys
import os
from pathlib import Path
import io

if sys.platform == 'win32':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

def inspect_file(file_path: str, max_rows: int = 50):
    """Проверяет структуру Excel файла"""
    print("=" * 70)
    print(f"ПРОВЕРКА СТРУКТУРЫ ФАЙЛА: {os.path.basename(file_path)}")
    print("=" * 70)
    print()
    
    file_ext = os.path.splitext(file_path)[1].lower()
    
    if file_ext == '.xls':
        try:
            import xlrd
            book = xlrd.open_workbook(file_path)
            sheet = book.sheet_by_index(0)
            
            print(f"Количество строк: {sheet.nrows}")
            print(f"Количество столбцов: {sheet.ncols}")
            print()
            print("Первые строки файла:")
            print("-" * 70)
            
            for row_idx in range(min(max_rows, sheet.nrows)):
                row_data = []
                for col_idx in range(min(10, sheet.ncols)):  # Первые 10 столбцов
                    cell_value = sheet.cell_value(row_idx, col_idx)
                    if isinstance(cell_value, float) and cell_value == int(cell_value):
                        cell_value = int(cell_value)
                    row_data.append(str(cell_value)[:30])  # Ограничиваем длину
                print(f"Строка {row_idx + 1}: {' | '.join(row_data)}")
        except Exception as e:
            print(f"Ошибка чтения .xls файла: {e}")
    else:
        try:
            from openpyxl import load_workbook
            wb = load_workbook(file_path, data_only=True)
            ws = wb.active
            
            print(f"Количество строк: {ws.max_row}")
            print(f"Количество столбцов: {ws.max_column}")
            print()
            print("Первые строки файла:")
            print("-" * 70)
            
            for row_idx in range(1, min(max_rows + 1, ws.max_row + 1)):
                row_data = []
                for col_idx in range(1, min(11, ws.max_column + 1)):  # Первые 10 столбцов
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell_value = cell.value
                    if cell_value is None:
                        cell_value = ""
                    else:
                        cell_value = str(cell_value)[:30]  # Ограничиваем длину
                    row_data.append(cell_value)
                print(f"Строка {row_idx}: {' | '.join(row_data)}")
        except Exception as e:
            print(f"Ошибка чтения .xlsx файла: {e}")
    
    print()
    print("=" * 70)

if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("Использование: python inspect_excel.py <путь_к_файлу>")
        print()
        print("Примеры:")
        print('  python inspect_excel.py "D:\\Excel file\\p-1 (28.08.25).xls"')
        print('  python inspect_excel.py "D:\\Excel file\\a-1 (28.08.25).xls"')
        sys.exit(1)
    
    file_path = sys.argv[1]
    if not os.path.exists(file_path):
        print(f"Файл не найден: {file_path}")
        sys.exit(1)
    
    inspect_file(file_path)

