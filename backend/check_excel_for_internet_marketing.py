"""Проверка как записан предмет Интернет-маркетинг в Excel"""
import os
from openpyxl import load_workbook
import sys
import io

# Исправление кодировки для Windows
if sys.platform == 'win32':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

excel_dir = r"D:\Excel file"
group_code = "S-4"

print("=" * 80)
print(f"ПОИСК ФАЙЛОВ ДЛЯ ГРУППЫ {group_code}")
print("=" * 80)

# Ищем файлы для группы S-4
excel_files = []
for filename in os.listdir(excel_dir):
    if group_code.lower() in filename.lower() and (filename.endswith('.xlsx') or filename.endswith('.xls')):
        excel_files.append(os.path.join(excel_dir, filename))

if not excel_files:
    print(f"❌ Файлы для группы {group_code} не найдены в {excel_dir}")
    exit(1)

print(f"\nНайдено файлов: {len(excel_files)}\n")

for filepath in excel_files:
    print(f"📄 Файл: {os.path.basename(filepath)}")
    try:
        wb = load_workbook(filepath, data_only=True)
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"\n  Лист: {sheet_name}")
            
            # Ищем ячейки с "интернет" или "маркетинг"
            found_cells = []
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        cell_lower = cell.value.lower()
                        if 'интернет' in cell_lower or 'маркетинг' in cell_lower:
                            found_cells.append((cell.row, cell.column, cell.value))
            
            if found_cells:
                print(f"  Найдено {len(found_cells)} ячеек с 'интернет' или 'маркетинг':")
                for row, col, value in found_cells[:10]:  # Показываем первые 10
                    value_short = value[:80] + "..." if len(value) > 80 else value
                    # Подсчитываем заглавные и строчные буквы
                    uppercase = sum(1 for c in value if c.isalpha() and c.isupper())
                    lowercase = sum(1 for c in value if c.isalpha() and c.islower())
                    total = uppercase + lowercase
                    upper_pct = (uppercase / total * 100) if total > 0 else 0
                    print(f"    Строка {row}, Колонка {col}: {value_short}")
                    print(f"      Заглавных: {uppercase}, Строчных: {lowercase}, Всего: {total}, % заглавных: {upper_pct:.1f}%")
            else:
                print("  Не найдено ячеек с 'интернет' или 'маркетинг'")
        
        wb.close()
    except Exception as e:
        print(f"  ❌ Ошибка при чтении файла: {e}")

