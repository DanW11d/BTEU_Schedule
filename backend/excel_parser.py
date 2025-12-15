"""
Парсер Excel файлов с расписанием для BTEU Schedule
Обрабатывает файлы с разделителями "─────────" и определяет четность недели
"""
import re
import os
from typing import List, Dict, Optional, Tuple, Any
from openpyxl import load_workbook
from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet


class ExcelScheduleParser:
    """Парсер Excel файлов с расписанием"""
    
    # Разделитель недель
    WEEK_SEPARATOR = "─────────"
    
    # Паттерны для определения четности недели
    ODD_WEEK_PATTERNS = ["ПОД ЧЕТЫ", "ПОДЧЕТЫ", "НЕЧЕТ", "НЕЧЁТ", "ODD"]
    EVEN_WEEK_PATTERNS = ["ОДЦ ЧЕТЫ", "ОДЦЧЕТЫ", "ЧЕТ", "ЧЁТ", "EVEN"]
    
    def __init__(self, file_path: str):
        """
        Инициализация парсера
        
        Args:
            file_path: Путь к Excel файлу
        """
        self.file_path = file_path
        self.workbook = None
        self.worksheet = None
        
    def load_file(self) -> None:
        """Загружает Excel файл (поддерживает .xlsx и .xls)"""
        try:
            file_ext = os.path.splitext(self.file_path)[1].lower()
            
            if file_ext == '.xls':
                # Для .xls используем xlrd и конвертируем в формат openpyxl
                try:
                    import xlrd
                    # Читаем .xls через xlrd
                    xls_book = xlrd.open_workbook(self.file_path)
                    xls_sheet = xls_book.sheet_by_index(0)
                    
                    # Создаем временный .xlsx файл
                    from openpyxl import Workbook
                    temp_wb = Workbook()
                    temp_ws = temp_wb.active
                    
                    # Копируем данные
                    for row_idx in range(xls_sheet.nrows):
                        for col_idx in range(xls_sheet.ncols):
                            cell_value = xls_sheet.cell_value(row_idx, col_idx)
                            temp_ws.cell(row=row_idx + 1, column=col_idx + 1, value=cell_value)
                    
                    self.workbook = temp_wb
                    self.worksheet = temp_ws
                except ImportError:
                    raise ValueError("Для чтения .xls файлов требуется библиотека xlrd. Установите: pip install xlrd")
                except Exception as e:
                    raise ValueError(f"Ошибка чтения .xls файла: {str(e)}")
            else:
                # Для .xlsx используем openpyxl
                self.workbook = load_workbook(self.file_path, data_only=True)
                # Берем первый лист
                self.worksheet = self.workbook.active
        except Exception as e:
            raise ValueError(f"Ошибка загрузки файла: {str(e)}")
    
    def parse(self, group_code: str, group_id: int) -> List[Dict]:
        """
        Парсит расписание из Excel файла
        
        Args:
            group_code: Код группы (например, "П-1")
            group_id: ID группы в БД
            
        Returns:
            Список словарей с данными занятий
        """
        if not self.worksheet:
            self.load_file()
        
        # Пытаемся определить формат файла
        if self._is_column_format():
            # Используем парсер для столбцового формата
            from excel_parser_v2 import ExcelScheduleParserV2
            parser_v2 = ExcelScheduleParserV2(self.file_path)
            parser_v2.workbook = self.workbook
            parser_v2.worksheet = self.worksheet
            return parser_v2.parse(group_code, group_id)
        
        # Используем старый построчный формат
        lessons = []
        
        # Находим начало данных (после шапки)
        start_row = self._find_data_start()
        
        # Находим конец данных (до подвала)
        end_row = self._find_data_end()
        
        # Парсим данные построчно
        current_week_parity = None  # "odd", "even", или None
        
        for row_idx in range(start_row, end_row + 1):
            row = list(self.worksheet.iter_rows(min_row=row_idx, max_row=row_idx, values_only=False))
            if not row or not row[0]:
                continue
            
            row_cells = row[0]
            
            # Проверяем на разделитель недель
            week_parity = self._detect_week_separator(row_cells)
            if week_parity:
                current_week_parity = week_parity
                continue
            
            # Парсим строку с занятием
            lesson = self._parse_lesson_row(row_cells, group_id, group_code, current_week_parity)
            if lesson:
                lessons.append(lesson)
        
        return lessons
    
    def _is_column_format(self) -> bool:
        """Определяет, является ли файл столбцовым форматом"""
        # Проверяем первые 20 строк на наличие заголовка "Дни | Время | ..."
        for row_idx in range(1, min(20, self.worksheet.max_row + 1)):
            row_str = " ".join([
                str(self.worksheet.cell(row=row_idx, column=col).value or "")
                for col in range(1, min(10, self.worksheet.max_column + 1))
            ]).lower()
            
            if "дни" in row_str and "время" in row_str:
                return True
        
        return False
    
    def _find_data_start(self) -> int:
        """
        Находит первую строку с данными (после шапки)
        
        Returns:
            Номер строки начала данных
        """
        # Обычно шапка занимает первые 2-5 строк
        # Ищем первую строку с номером пары (1, 2, 3...) или днем недели
        for row_idx in range(1, min(10, self.worksheet.max_row + 1)):
            row = list(self.worksheet.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))
            if not row or not row[0]:
                continue
            
            first_cell = str(row[0][0] or "").strip()
            
            # Проверяем, является ли это началом данных
            if self._is_data_row(first_cell):
                return row_idx
        
        # Если не нашли, начинаем с 3-й строки
        return 3
    
    def _find_data_end(self) -> int:
        """
        Находит последнюю строку с данными (до подвала)
        
        Returns:
            Номер строки конца данных
        """
        # Ищем последнюю строку с данными
        # Обычно подвал начинается с пустых строк или служебной информации
        max_row = self.worksheet.max_row
        
        for row_idx in range(max_row, max(1, max_row - 20), -1):
            row = list(self.worksheet.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))
            if not row or not row[0]:
                continue
            
            # Проверяем, есть ли в строке данные
            has_data = any(cell and str(cell).strip() for cell in row[0] if cell is not None)
            if has_data and not self._is_footer_row(row[0]):
                return row_idx
        
        return max_row
    
    def _is_data_row(self, first_cell: str) -> bool:
        """Проверяет, является ли строка строкой с данными"""
        if not first_cell:
            return False
        
        # Проверяем на номер пары
        if first_cell.isdigit() and 1 <= int(first_cell) <= 7:
            return True
        
        # Проверяем на день недели
        days = ["понедельник", "вторник", "среда", "четверг", "пятница", "суббота"]
        if first_cell.lower() in days:
            return True
        
        return False
    
    def _is_footer_row(self, row: Tuple) -> bool:
        """Проверяет, является ли строка подвалом"""
        row_str = " ".join(str(cell or "") for cell in row).lower()
        
        # Паттерны подвала
        footer_patterns = [
            "итого", "всего", "подпись", "дата", "утверждено",
            "расписание", "составлено", "проверено"
        ]
        
        return any(pattern in row_str for pattern in footer_patterns)
    
    def _detect_week_separator(self, row_cells: List[Cell]) -> Optional[str]:
        """
        Определяет разделитель недель и возвращает четность
        
        Returns:
            "odd", "even" или None
        """
        # Объединяем все ячейки строки в одну строку
        row_text = " ".join(
            str(cell.value or "").strip() 
            for cell in row_cells 
            if cell and cell.value
        ).upper()
        
        # Проверяем на разделитель
        if self.WEEK_SEPARATOR not in row_text and "─" * 5 not in row_text:
            return None
        
        # Определяем четность по паттернам
        for pattern in self.ODD_WEEK_PATTERNS:
            if pattern in row_text:
                return "odd"
        
        for pattern in self.EVEN_WEEK_PATTERNS:
            if pattern in row_text:
                return "even"
        
        # Если разделитель есть, но четность не определена, возвращаем None
        # (будет использоваться текущая четность)
        return None
    
    def _parse_lesson_row(
        self, 
        row_cells: List[Cell], 
        group_id: int, 
        group_code: str,
        week_parity: Optional[str]
    ) -> Optional[Dict]:
        """
        Парсит строку с занятием
        
        Args:
            row_cells: Список ячеек строки
            group_id: ID группы
            group_code: Код группы
            week_parity: Четность недели ("odd", "even", или None для "both")
            
        Returns:
            Словарь с данными занятия или None
        """
        # Преобразуем ячейки в значения
        values = [str(cell.value or "").strip() if cell and cell.value else "" for cell in row_cells]
        
        # Пропускаем пустые строки
        if not any(values):
            return None
        
        # Пропускаем строки-разделители
        row_text = " ".join(values).upper()
        if self.WEEK_SEPARATOR in row_text or "─" * 5 in row_text:
            return None
        
        # Типичная структура строки:
        # [Номер пары, День недели, Предмет, Преподаватель, Аудитория, ...]
        # Или: [День недели, Номер пары, Предмет, Преподаватель, Аудитория, ...]
        
        lesson_number = None
        day_of_week = None
        subject = None
        teacher = None
        classroom = None
        
        # Пытаемся найти номер пары (1-7)
        for i, val in enumerate(values):
            if val.isdigit() and 1 <= int(val) <= 7:
                lesson_number = int(val)
                # День недели обычно рядом
                if i > 0:
                    day_of_week = self._parse_day_of_week(values[i - 1])
                if not day_of_week and i < len(values) - 1:
                    day_of_week = self._parse_day_of_week(values[i + 1])
                break
        
        # Если не нашли номер пары, пытаемся найти день недели
        if not lesson_number:
            for i, val in enumerate(values):
                day = self._parse_day_of_week(val)
                if day:
                    day_of_week = day
                    # Номер пары может быть рядом
                    if i > 0 and values[i - 1].isdigit():
                        lesson_number = int(values[i - 1])
                    elif i < len(values) - 1 and values[i + 1].isdigit():
                        lesson_number = int(values[i + 1])
                    break
        
        # Если не нашли номер пары или день, пропускаем строку
        if not lesson_number or not day_of_week:
            return None
        
        # Защита от некорректных значений lesson_number
        if lesson_number < 1:
            lesson_number = 1
        if lesson_number > 7:
            lesson_number = 7
        
        # Ищем предмет (обычно самая длинная строка с текстом)
        subject_candidates = [v for v in values if v and len(v) > 5 and not v.isdigit()]
        if subject_candidates:
            subject = subject_candidates[0]
            # Определяем тип занятия по регистру
            lesson_type = self._detect_lesson_type(subject)
        else:
            return None
        
        # Ищем преподавателя (обычно содержит "доц.", "проф.", инициалы)
        teacher_pattern = re.compile(r'(доц|проф|ст\.|преп)\.?\s*[А-ЯЁ]\.?\s*[А-ЯЁ]\.?', re.IGNORECASE)
        for val in values:
            if teacher_pattern.search(val):
                teacher = val
                break
        
        # Ищем аудиторию (обычно формат "3-11", "5-41", или просто номер)
        classroom_pattern = re.compile(r'\d+[-–]\d+|\d+[а-яё]?', re.IGNORECASE)
        for val in values:
            if classroom_pattern.match(val) and val != str(lesson_number):
                classroom = val
                break
        
        # Если не нашли аудиторию, но есть еще значения, берем последнее
        if not classroom and len(values) > 3:
            for val in reversed(values):
                if val and val != subject and val != teacher and not val.isdigit():
                    classroom = val
                    break
        
        # Определяем четность недели
        if not week_parity:
            week_parity = "both"  # По умолчанию каждую неделю
        
        return {
            'group_id': group_id,
            'group_code': group_code,
            'day_of_week': day_of_week,
            'lesson_number': lesson_number,
            'subject': subject,
            'teacher': teacher or None,
            'classroom': classroom or None,
            'lesson_type': lesson_type,
            'week_parity': week_parity,
            'building': None,  # Можно извлечь из аудитории позже
            'notes': None
        }
    
    def _parse_day_of_week(self, text: str) -> Optional[int]:
        """
        Парсит день недели из текста
        
        Returns:
            1-6 (Понедельник-Суббота) или None
        """
        if not text:
            return None
        
        text_lower = text.lower().strip()
        
        days_map = {
            'понедельник': 1, 'пн': 1, 'monday': 1, 'mon': 1,
            'вторник': 2, 'вт': 2, 'tuesday': 2, 'tue': 2,
            'среда': 3, 'ср': 3, 'wednesday': 3, 'wed': 3,
            'четверг': 4, 'чт': 4, 'thursday': 4, 'thu': 4,
            'пятница': 5, 'пт': 5, 'friday': 5, 'fri': 5,
            'суббота': 6, 'сб': 6, 'saturday': 6, 'sat': 6
        }
        
        return days_map.get(text_lower)
    
    def _detect_lesson_type(self, subject: str) -> str:
        """
        Определяет тип занятия по регистру букв в названии предмета
        
        Args:
            subject: Название предмета
            
        Returns:
            "lecture", "practice", или "laboratory"
        """
        if not subject:
            return "practice"
        
        # Подсчитываем заглавные и строчные буквы
        uppercase_count = sum(1 for c in subject if c.isupper() and c.isalpha())
        lowercase_count = sum(1 for c in subject if c.islower() and c.isalpha())
        total_letters = uppercase_count + lowercase_count
        
        if total_letters == 0:
            return "practice"
        
        # Если больше 50% заглавных букв - это лекция
        uppercase_ratio = uppercase_count / total_letters
        
        if uppercase_ratio > 0.5:
            return "lecture"
        elif "лаб" in subject.lower() or "лабораторн" in subject.lower():
            return "laboratory"
        else:
            return "practice"
    
    def validate_lessons(self, lessons: List[Dict]) -> Tuple[bool, List[str]]:
        """
        Валидирует список занятий
        
        Args:
            lessons: Список словарей с данными занятий
            
        Returns:
            (is_valid, errors)
        """
        errors = []
        
        for i, lesson in enumerate(lessons, 1):
            # Проверка обязательных полей
            if not lesson.get('group_id'):
                errors.append(f"Занятие {i}: отсутствует group_id")
            if not lesson.get('day_of_week') or not (1 <= lesson['day_of_week'] <= 6):
                errors.append(f"Занятие {i}: некорректный day_of_week ({lesson.get('day_of_week')})")
            if not lesson.get('lesson_number') or not (1 <= lesson['lesson_number'] <= 7):
                errors.append(f"Занятие {i}: некорректный lesson_number ({lesson.get('lesson_number')})")
            if not lesson.get('subject'):
                errors.append(f"Занятие {i}: отсутствует subject")
            if lesson.get('lesson_type') not in ['lecture', 'practice', 'laboratory']:
                errors.append(f"Занятие {i}: некорректный lesson_type ({lesson.get('lesson_type')})")
            if lesson.get('week_parity') not in ['odd', 'even', 'both']:
                errors.append(f"Занятие {i}: некорректный week_parity ({lesson.get('week_parity')})")
        
        return len(errors) == 0, errors

